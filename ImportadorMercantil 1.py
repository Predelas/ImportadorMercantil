from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import date
from tkinter.filedialog import askopenfilename
import time

base = [
    'NUM_BANCO', 'NOM_BANCO', 'NUM_PROPOSTA', 'NUM_CONTRATO',
    'DSC_TIPO_PROPOSTA_EMPRESTIMO', 'COD_PRODUTO', 'DSC_PRODUTO',
    'DAT_CTR_INCLUSAO', 'DSC_SITUACAO_EMPRESTIMO', 'DAT_EMPRESTIMO',
    'COD_EMPREGADOR', 'DSC_CONVENIO', 'COD_ORGAO', 'NOM_ORGAO',
    'COD_PRODUTOR_VENDA', 'NOM_PRODUTOR_VENDA', 'NIC_CTR_USUARIO',
    'COD_CPF_CLIENTE', 'NOM_CLIENTE', 'DAT_NASCIMENTO', 'NUM_IDENTIDADE',
    'NOM_LOGRADOURO', 'NUM_PREDIO', 'DSC_CMPLMNT_ENDRC', 'NOM_BAIRRO',
    'NOM_LOCALIDADE', 'SIG_UNIDADE_FEDERACAO', 'COD_ENDRCMNT_PSTL',
    'NUM_TELEFONE', 'NUM_TELEFONE_CELULAR', 'NOM_MAE', 'NOM_PAI',
    'NUM_BENEFICIO', 'QTD_PARCELA', 'VAL_PRESTACAO', 'VAL_BRUTO',
    'VAL_SALDO_RECOMPRA', 'VAL_SALDO_REFINANCIAMENTO', 'VAL_LIQUIDO',
    'PCR_PMT_PAGO_REF', 'DAT_CREDITO', 'DAT_CONFIRMACAO', 'VAL_REPASSE',
    'PCL_COMISSAO', 'VAL_COMISSAO', 'COD_UNIDADE_EMPRESA',
    'COD_SITUACAO_EMPRESTIMO', 'DAT_ESTORNO', 'DSC_OBSERVACAO',
    'NUM_CPF_AGENTE', 'NUM_OBJETO_ECT', 'PCL_TAXA_EMPRESTIMO',
    'DSC_TIPO_FORMULARIO_EMPRESTIMO', 'DSC_TIPO_CREDITO_EMPRESTIMO',
    'NOM_GRUPO_UNIDADE_EMPRESA', 'COD_PROPOSTA_EMPRESTIMO',
    'COD_GRUPO_UNIDADE_EMPRESA', 'COD_TIPO_FUNCAO',
    'COD_TIPO_PROPOSTA_EMPRESTIMO', 'COD_LOJA_DIGITACAO', 'VAL_SEGURO'
]

arquivo_excel = Workbook()
planilha1 = arquivo_excel.active
planilha1.title = ("Produção Novo")
path = askopenfilename()
relatorio = load_workbook(path)
planilha2 = relatorio.active
planilha1.append(base)
arquivo_excel.save("WORKBANK_MODELO_INTEGRACAO Mercantil " + str(date.today()) + ".xlsx")

aux = planilha1.cell(row=999, column=999).value
max_linha = planilha2.max_row + 1

for i in range(2, max_linha):

		op = str(planilha2.cell(row=i, column=5).value).split()
	
		planilha1.cell(row=i, column=1, value="389")
		planilha1.cell(row=i, column=2, value="BANCO MERCANTIL DO BRASIL")
		planilha1.cell(row=i, column=3, value=(planilha2.cell(row=i, column=1).value))
		planilha1.cell(row=i, column=4, value=(planilha2.cell(row=i, column=1).value))
		
		planilha1.cell(row=i, column=6, value=(planilha2.cell(row=i, column=4).value))
		planilha1.cell(row=i, column=8, value=(planilha2.cell(row=i, column=13).value))
		planilha1.cell(row=i, column=9, value=(planilha2.cell(row=i, column=16).value))
		planilha1.cell(row=i, column=10, value=(planilha2.cell(row=i, column=13).value))
		planilha1.cell(row=i, column=17, value=(planilha2.cell(row=i, column=12).value))
		planilha1.cell(row=i, column=18, value=(planilha2.cell(row=i, column=36).value))
		planilha1.cell(row=i, column=19, value=(planilha2.cell(row=i, column=37).value))
		planilha1.cell(row=i, column=34, value=(planilha2.cell(row=i, column=19).value))
		planilha1.cell(row=i, column=35, value=(planilha2.cell(row=i, column=20).value))
		planilha1.cell(row=i, column=36, value=(planilha2.cell(row=i, column=21).value))
		planilha1.cell(row=i, column=39, value=(planilha2.cell(row=i, column=22).value))
		planilha1.cell(row=i, column=41, value=(planilha2.cell(row=i, column=58).value))
		planilha1.cell(row=i, column=53, value="DIGITAL")	

		if "FGTS" in op:

				planilha1.cell(row=i, column=7, value=("FGTS - NOVO"))
				planilha1.cell(row=i, column=12, value=("FGTS"))
				planilha1.cell(row=i, column=5, value=("NOVO"))

		else:
				planilha1.cell(row=i, column=7, value=(planilha2.cell(row=i, column=5).value))
				planilha1.cell(row=i, column=5, value=(planilha2.cell(row=i, column=14).value))

				if "INSS" in op:
						planilha1.cell(row=i, column=12, value=("INSS"))
				else:
						arquivo_excel2 = Workbook()
						planilhaE = arquivo_excel2.active
						planilhaE.title = ("Erros")
						planilhaE.append(op)
						arquivo_excel2.save("Erros.xlsx")
		if str(planilha2.cell(row=i, column=14).value) == "SaqueAniversarioFgts":
				planilha1.cell(row=i, column=7, value=("FGTS - NOVO"))
				planilha1.cell(row=i, column=12, value=("FGTS"))
				planilha1.cell(row=i, column=5, value=("NOVO"))

arquivo_excel.save("WORKBANK_MODELO_INTEGRACAO Mercantil " + str(date.today()) + ".xlsx")
